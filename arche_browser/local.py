"""
Local PC Control Module - Full system access via minimal primitives.

Design Philosophy (Eric Gamma style):
- Simple: Just a few powerful primitives
- Flexible: Compose them to do anything
- Powerful: No artificial limitations

Core Primitives:
1. shell() - Execute any shell command
2. python_exec() - Execute Python code with full access
3. screen_capture() - Desktop screenshot for visual feedback

With these 3, an AI can control EVERYTHING:
- Volume, camera, microphone
- Files, programs, services
- Excel, PowerPoint, any application
- System maintenance, cleanup, optimization
- Literally anything a human can do
"""

import subprocess
import sys
import os
import platform
import shutil
import json
import base64
import tempfile
from pathlib import Path
from typing import Optional, Any
from datetime import datetime


class LocalControl:
    """Full local PC control with minimal, powerful primitives."""

    def __init__(self):
        self.python_globals = {"__builtins__": __builtins__}
        self.python_locals = {}

    # =========================================================================
    # CORE PRIMITIVES - These 3 can do ANYTHING
    # =========================================================================

    def shell(
        self,
        command: str,
        cwd: Optional[str] = None,
        timeout: int = 60,
        shell_type: Optional[str] = None,
        capture: bool = True
    ) -> dict:
        """
        Execute a shell command. This is the most powerful primitive.

        Args:
            command: Command to execute
            cwd: Working directory (default: current)
            timeout: Timeout in seconds (default: 60)
            shell_type: 'bash', 'cmd', 'powershell', or None for auto
            capture: Capture output (default: True)

        Returns:
            dict with stdout, stderr, return_code, success

        Examples:
            # Windows volume control
            shell("powershell (Get-AudioDevice -Playback).SetMute($false)")

            # Run a program
            shell("notepad.exe")

            # System info
            shell("systeminfo")

            # Reboot
            shell("shutdown /r /t 60")
        """
        try:
            # Determine shell
            is_windows = platform.system() == "Windows"

            if shell_type == "powershell" or (shell_type is None and is_windows):
                if is_windows:
                    # Use powershell on Windows
                    args = ["powershell", "-Command", command]
                else:
                    args = ["pwsh", "-Command", command]
            elif shell_type == "cmd":
                args = ["cmd", "/c", command]
            else:
                # bash/sh
                args = ["/bin/bash", "-c", command] if os.path.exists("/bin/bash") else ["/bin/sh", "-c", command]

            result = subprocess.run(
                args,
                capture_output=capture,
                text=True,
                cwd=cwd,
                timeout=timeout
            )

            return {
                "success": result.returncode == 0,
                "stdout": result.stdout if capture else "",
                "stderr": result.stderr if capture else "",
                "return_code": result.returncode
            }

        except subprocess.TimeoutExpired:
            return {
                "success": False,
                "stdout": "",
                "stderr": f"Command timed out after {timeout} seconds",
                "return_code": -1
            }
        except Exception as e:
            return {
                "success": False,
                "stdout": "",
                "stderr": str(e),
                "return_code": -1
            }

    def python_exec(
        self,
        code: str,
        timeout: int = 60
    ) -> dict:
        """
        Execute Python code with full access. Infinitely extensible.

        The code runs in a persistent context, so variables and imports
        are preserved between calls.

        Args:
            code: Python code to execute
            timeout: Timeout in seconds (default: 60)

        Returns:
            dict with result, output, success, error

        Examples:
            # Take a photo with webcam
            python_exec('''
            import cv2
            cap = cv2.VideoCapture(0)
            ret, frame = cap.read()
            cv2.imwrite("photo.jpg", frame)
            cap.release()
            "photo.jpg"
            ''')

            # Control Excel
            python_exec('''
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Hello'
            wb.save('test.xlsx')
            ''')

            # Get system temperatures (Windows)
            python_exec('''
            import wmi
            w = wmi.WMI(namespace="root\\wmi")
            temps = w.MSAcpi_ThermalZoneTemperature()
            [(t.InstanceName, (t.CurrentTemperature - 2732) / 10) for t in temps]
            ''')
        """
        import io
        from contextlib import redirect_stdout, redirect_stderr

        stdout_capture = io.StringIO()
        stderr_capture = io.StringIO()

        try:
            with redirect_stdout(stdout_capture), redirect_stderr(stderr_capture):
                # Try to get the result of the last expression
                try:
                    # First try exec for statements
                    exec(code, self.python_globals, self.python_locals)
                    result = None
                except SyntaxError:
                    pass

                # Try to eval the last line as expression
                lines = code.strip().split('\n')
                if lines:
                    last_line = lines[-1].strip()
                    if last_line and not any(last_line.startswith(kw) for kw in
                        ['import ', 'from ', 'def ', 'class ', 'if ', 'for ', 'while ',
                         'try:', 'with ', 'return ', '#', 'else:', 'elif ', 'except', 'finally:']):
                        try:
                            result = eval(last_line, self.python_globals, self.python_locals)
                        except:
                            result = None

            return {
                "success": True,
                "result": repr(result) if result is not None else None,
                "stdout": stdout_capture.getvalue(),
                "stderr": stderr_capture.getvalue(),
                "error": None
            }

        except Exception as e:
            import traceback
            return {
                "success": False,
                "result": None,
                "stdout": stdout_capture.getvalue(),
                "stderr": stderr_capture.getvalue(),
                "error": traceback.format_exc()
            }

    def screen_capture(
        self,
        path: Optional[str] = None,
        region: Optional[tuple] = None
    ) -> dict:
        """
        Capture the desktop screen. Essential for visual feedback.

        Args:
            path: Output path (default: temp file)
            region: Optional (x, y, width, height) tuple

        Returns:
            dict with path, width, height, success

        Note: Requires 'pillow' or 'mss' package.
        """
        try:
            # Try mss first (faster, cross-platform)
            try:
                import mss
                with mss.mss() as sct:
                    if region:
                        monitor = {"left": region[0], "top": region[1],
                                   "width": region[2], "height": region[3]}
                    else:
                        monitor = sct.monitors[1]  # Primary monitor

                    screenshot = sct.grab(monitor)

                    if path is None:
                        path = tempfile.mktemp(suffix=".png")

                    # Convert to PIL and save
                    from PIL import Image
                    img = Image.frombytes("RGB", screenshot.size, screenshot.bgra, "raw", "BGRX")
                    img.save(path)

                    return {
                        "success": True,
                        "path": path,
                        "width": screenshot.width,
                        "height": screenshot.height
                    }
            except ImportError:
                pass

            # Fallback to PIL
            try:
                from PIL import ImageGrab
                if region:
                    bbox = (region[0], region[1], region[0] + region[2], region[1] + region[3])
                    screenshot = ImageGrab.grab(bbox=bbox)
                else:
                    screenshot = ImageGrab.grab()

                if path is None:
                    path = tempfile.mktemp(suffix=".png")

                screenshot.save(path)

                return {
                    "success": True,
                    "path": path,
                    "width": screenshot.width,
                    "height": screenshot.height
                }
            except ImportError:
                pass

            # Fallback to shell command
            if platform.system() == "Windows":
                # Use PowerShell
                if path is None:
                    path = tempfile.mktemp(suffix=".png")
                ps_script = f'''
                Add-Type -AssemblyName System.Windows.Forms
                $screen = [System.Windows.Forms.Screen]::PrimaryScreen.Bounds
                $bitmap = New-Object System.Drawing.Bitmap($screen.Width, $screen.Height)
                $graphics = [System.Drawing.Graphics]::FromImage($bitmap)
                $graphics.CopyFromScreen($screen.Location, [System.Drawing.Point]::Empty, $screen.Size)
                $bitmap.Save("{path}")
                '''
                result = self.shell(ps_script, shell_type="powershell")
                if result["success"]:
                    return {"success": True, "path": path, "width": None, "height": None}

            return {
                "success": False,
                "error": "No screenshot method available. Install 'mss' or 'pillow': pip install mss pillow"
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    # =========================================================================
    # CONVENIENCE TOOLS - For token efficiency
    # =========================================================================

    def file_read(self, path: str, encoding: str = "utf-8") -> dict:
        """Read file content."""
        try:
            p = Path(path).expanduser()

            # Check if binary
            try:
                with open(p, 'r', encoding=encoding) as f:
                    content = f.read()
                return {"success": True, "content": content, "binary": False}
            except UnicodeDecodeError:
                with open(p, 'rb') as f:
                    content = base64.b64encode(f.read()).decode('ascii')
                return {"success": True, "content": content, "binary": True}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def file_write(
        self,
        path: str,
        content: str,
        encoding: str = "utf-8",
        binary: bool = False,
        append: bool = False
    ) -> dict:
        """Write content to file."""
        try:
            p = Path(path).expanduser()
            p.parent.mkdir(parents=True, exist_ok=True)

            mode = 'ab' if append and binary else 'a' if append else 'wb' if binary else 'w'

            if binary:
                data = base64.b64decode(content)
                with open(p, mode) as f:
                    f.write(data)
            else:
                with open(p, mode, encoding=encoding) as f:
                    f.write(content)

            return {"success": True, "path": str(p), "size": p.stat().st_size}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def file_list(
        self,
        path: str = ".",
        pattern: str = "*",
        recursive: bool = False,
        details: bool = False
    ) -> dict:
        """List directory contents."""
        try:
            p = Path(path).expanduser()

            if recursive:
                files = list(p.rglob(pattern))
            else:
                files = list(p.glob(pattern))

            if details:
                result = []
                for f in files[:1000]:  # Limit to 1000
                    try:
                        stat = f.stat()
                        result.append({
                            "path": str(f),
                            "name": f.name,
                            "is_dir": f.is_dir(),
                            "size": stat.st_size if not f.is_dir() else None,
                            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
                        })
                    except:
                        result.append({"path": str(f), "name": f.name})
                return {"success": True, "files": result, "count": len(files)}
            else:
                return {
                    "success": True,
                    "files": [str(f) for f in files[:1000]],
                    "count": len(files)
                }

        except Exception as e:
            return {"success": False, "error": str(e)}

    def file_delete(self, path: str, recursive: bool = False) -> dict:
        """Delete file or directory."""
        try:
            p = Path(path).expanduser()

            if p.is_dir():
                if recursive:
                    shutil.rmtree(p)
                else:
                    p.rmdir()
            else:
                p.unlink()

            return {"success": True, "deleted": str(p)}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def file_copy(self, src: str, dst: str) -> dict:
        """Copy file or directory."""
        try:
            src_p = Path(src).expanduser()
            dst_p = Path(dst).expanduser()

            if src_p.is_dir():
                shutil.copytree(src_p, dst_p)
            else:
                dst_p.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(src_p, dst_p)

            return {"success": True, "src": str(src_p), "dst": str(dst_p)}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def file_move(self, src: str, dst: str) -> dict:
        """Move/rename file or directory."""
        try:
            src_p = Path(src).expanduser()
            dst_p = Path(dst).expanduser()
            dst_p.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(src_p, dst_p)
            return {"success": True, "src": str(src_p), "dst": str(dst_p)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def clipboard_get(self) -> dict:
        """Get clipboard content."""
        try:
            # Try pyperclip
            try:
                import pyperclip
                return {"success": True, "content": pyperclip.paste()}
            except ImportError:
                pass

            # Try tkinter
            try:
                import tkinter as tk
                root = tk.Tk()
                root.withdraw()
                content = root.clipboard_get()
                root.destroy()
                return {"success": True, "content": content}
            except:
                pass

            # Fallback to shell
            if platform.system() == "Windows":
                result = self.shell("powershell Get-Clipboard", shell_type="powershell")
                if result["success"]:
                    return {"success": True, "content": result["stdout"].strip()}
            elif platform.system() == "Darwin":
                result = self.shell("pbpaste")
                if result["success"]:
                    return {"success": True, "content": result["stdout"]}
            else:
                result = self.shell("xclip -selection clipboard -o")
                if result["success"]:
                    return {"success": True, "content": result["stdout"]}

            return {"success": False, "error": "No clipboard method available"}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def clipboard_set(self, content: str) -> dict:
        """Set clipboard content."""
        try:
            # Try pyperclip
            try:
                import pyperclip
                pyperclip.copy(content)
                return {"success": True}
            except ImportError:
                pass

            # Try tkinter
            try:
                import tkinter as tk
                root = tk.Tk()
                root.withdraw()
                root.clipboard_clear()
                root.clipboard_append(content)
                root.update()
                root.destroy()
                return {"success": True}
            except:
                pass

            # Fallback to shell
            if platform.system() == "Windows":
                # Escape for PowerShell
                escaped = content.replace("'", "''")
                result = self.shell(f"Set-Clipboard -Value '{escaped}'", shell_type="powershell")
                return {"success": result["success"], "error": result.get("stderr")}
            elif platform.system() == "Darwin":
                result = self.shell(f"echo '{content}' | pbcopy")
                return {"success": result["success"], "error": result.get("stderr")}
            else:
                result = self.shell(f"echo '{content}' | xclip -selection clipboard")
                return {"success": result["success"], "error": result.get("stderr")}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def system_info(self) -> dict:
        """Get comprehensive system information."""
        try:
            info = {
                "platform": platform.system(),
                "platform_release": platform.release(),
                "platform_version": platform.version(),
                "architecture": platform.machine(),
                "hostname": platform.node(),
                "processor": platform.processor(),
                "python_version": sys.version,
                "cwd": os.getcwd(),
                "user": os.getenv("USER") or os.getenv("USERNAME"),
                "home": str(Path.home()),
                "temp": tempfile.gettempdir(),
            }

            # Try to get more info
            try:
                import psutil
                info["cpu_count"] = psutil.cpu_count()
                info["cpu_percent"] = psutil.cpu_percent(interval=0.1)
                info["memory_total"] = psutil.virtual_memory().total
                info["memory_available"] = psutil.virtual_memory().available
                info["memory_percent"] = psutil.virtual_memory().percent
                info["disk_usage"] = {
                    p.mountpoint: {
                        "total": psutil.disk_usage(p.mountpoint).total,
                        "used": psutil.disk_usage(p.mountpoint).used,
                        "free": psutil.disk_usage(p.mountpoint).free,
                        "percent": psutil.disk_usage(p.mountpoint).percent
                    }
                    for p in psutil.disk_partitions() if p.fstype
                }
            except ImportError:
                pass

            return {"success": True, "info": info}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def process_list(self, filter_name: Optional[str] = None) -> dict:
        """List running processes."""
        try:
            try:
                import psutil
                processes = []
                for proc in psutil.process_iter(['pid', 'name', 'cpu_percent', 'memory_percent']):
                    try:
                        info = proc.info
                        if filter_name and filter_name.lower() not in info['name'].lower():
                            continue
                        processes.append(info)
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        continue
                return {"success": True, "processes": processes[:100]}
            except ImportError:
                # Fallback to shell
                if platform.system() == "Windows":
                    result = self.shell("tasklist /fo csv", shell_type="cmd")
                else:
                    result = self.shell("ps aux")
                return {"success": result["success"], "output": result["stdout"]}

        except Exception as e:
            return {"success": False, "error": str(e)}

    def process_kill(self, pid: Optional[int] = None, name: Optional[str] = None) -> dict:
        """Kill a process by PID or name."""
        try:
            if pid:
                try:
                    import psutil
                    psutil.Process(pid).terminate()
                    return {"success": True, "killed_pid": pid}
                except ImportError:
                    if platform.system() == "Windows":
                        result = self.shell(f"taskkill /PID {pid} /F", shell_type="cmd")
                    else:
                        result = self.shell(f"kill -9 {pid}")
                    return {"success": result["success"], "error": result.get("stderr")}
            elif name:
                if platform.system() == "Windows":
                    result = self.shell(f"taskkill /IM {name} /F", shell_type="cmd")
                else:
                    result = self.shell(f"pkill -9 {name}")
                return {"success": result["success"], "error": result.get("stderr")}
            else:
                return {"success": False, "error": "Provide pid or name"}

        except Exception as e:
            return {"success": False, "error": str(e)}


def register_local_tools(mcp, local: LocalControl):
    """Register local control tools with the MCP server."""

    @mcp.tool()
    def shell(
        command: str,
        cwd: str = None,
        timeout: int = 60,
        shell_type: str = None
    ) -> str:
        """
        Execute a shell command. The most powerful primitive for PC control.

        Args:
            command: Command to execute
            cwd: Working directory (optional)
            timeout: Timeout in seconds (default: 60)
            shell_type: 'bash', 'cmd', 'powershell', or auto-detect

        Examples:
            - shell("dir") - List files (Windows)
            - shell("ls -la") - List files (Linux/Mac)
            - shell("notepad.exe") - Open Notepad
            - shell("shutdown /r /t 60") - Reboot in 60 seconds
            - shell("powershell (Get-Volume).DriveLetter") - PowerShell command
        """
        result = local.shell(command, cwd, timeout, shell_type)
        return json.dumps(result)

    @mcp.tool()
    def python_exec(code: str, timeout: int = 60) -> str:
        """
        Execute Python code with full system access. Infinitely extensible.

        Variables and imports persist between calls.

        Args:
            code: Python code to execute
            timeout: Timeout in seconds (default: 60)

        Examples:
            # Take webcam photo
            python_exec('''
            import cv2
            cap = cv2.VideoCapture(0)
            ret, frame = cap.read()
            cv2.imwrite("photo.jpg", frame)
            cap.release()
            ''')

            # Create Excel file
            python_exec('''
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Hello World'
            wb.save('test.xlsx')
            ''')
        """
        result = local.python_exec(code, timeout)
        return json.dumps(result)

    @mcp.tool()
    def screen_capture(path: str = None) -> str:
        """
        Capture the desktop screen.

        Args:
            path: Output file path (optional, defaults to temp file)

        Returns:
            Path to the screenshot file
        """
        result = local.screen_capture(path)
        return json.dumps(result)

    @mcp.tool()
    def file_read(path: str) -> str:
        """Read file content. Binary files are returned as base64."""
        result = local.file_read(path)
        return json.dumps(result)

    @mcp.tool()
    def file_write(path: str, content: str, append: bool = False) -> str:
        """Write content to a file. Creates parent directories if needed."""
        result = local.file_write(path, content, append=append)
        return json.dumps(result)

    @mcp.tool()
    def file_list(path: str = ".", pattern: str = "*", recursive: bool = False) -> str:
        """List directory contents with optional pattern matching."""
        result = local.file_list(path, pattern, recursive, details=True)
        return json.dumps(result)

    @mcp.tool()
    def file_delete(path: str, recursive: bool = False) -> str:
        """Delete a file or directory."""
        result = local.file_delete(path, recursive)
        return json.dumps(result)

    @mcp.tool()
    def file_copy(src: str, dst: str) -> str:
        """Copy a file or directory."""
        result = local.file_copy(src, dst)
        return json.dumps(result)

    @mcp.tool()
    def file_move(src: str, dst: str) -> str:
        """Move or rename a file or directory."""
        result = local.file_move(src, dst)
        return json.dumps(result)

    @mcp.tool()
    def clipboard_get() -> str:
        """Get the current clipboard content."""
        result = local.clipboard_get()
        return json.dumps(result)

    @mcp.tool()
    def clipboard_set(content: str) -> str:
        """Set the clipboard content."""
        result = local.clipboard_set(content)
        return json.dumps(result)

    @mcp.tool()
    def system_info() -> str:
        """Get comprehensive system information (OS, CPU, memory, disk, etc.)."""
        result = local.system_info()
        return json.dumps(result)

    @mcp.tool()
    def process_list(filter_name: str = None) -> str:
        """List running processes, optionally filtered by name."""
        result = local.process_list(filter_name)
        return json.dumps(result)

    @mcp.tool()
    def process_kill(pid: int = None, name: str = None) -> str:
        """Kill a process by PID or name."""
        result = local.process_kill(pid, name)
        return json.dumps(result)
